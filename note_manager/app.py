from flask import Flask, render_template, redirect, url_for, request, flash
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from wtforms import StringField, FloatField
from wtforms.validators import InputRequired
from fpdf import FPDF
import openpyxl
from flask_migrate import Migrate

# Créez l'application Flask
app = Flask(__name__)

# Configuration de la base de données
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///notes.db'
app.config['SECRET_KEY'] = 'votre_cle_secrete'

# Initialiser SQLAlchemy
db = SQLAlchemy(app)

# Initialiser Flask-Migrate
migrate = Migrate(app, db)

# Modèle pour les notes
class Note(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    titre = db.Column(db.String(100), nullable=False)
    note = db.Column(db.Float, nullable=False)

    def __repr__(self):
        return f'<Note {self.titre} - {self.note}>'

# Formulaire pour la gestion des notes
class NoteForm(FlaskForm):
    titre = StringField('Titre', validators=[InputRequired()])
    note = FloatField('Note', validators=[InputRequired()])

# Route pour la page d'accueil
@app.route('/')
def index():
    return render_template('index.html')

# Route pour afficher et gérer les notes
@app.route('/gestion', methods=['GET', 'POST'])
def gestion_notes():
    form = NoteForm()
    if form.validate_on_submit():
        new_note = Note(
            titre=form.titre.data,
            note=form.note.data
        )
        db.session.add(new_note)
        db.session.commit()
        flash("Note ajoutée avec succès", 'success')
        return redirect(url_for('gestion_notes'))

    notes = Note.query.all()
    return render_template('gestion_notes.html', form=form, notes=notes)

# Route pour afficher les notes avec pagination
@app.route('/afficher')
def afficher_notes():
    page = request.args.get('page', 1, type=int)
    notes = Note.query.paginate(page=page, per_page=5)  # Affiche 5 notes par page
    return render_template('afficher_notes.html', notes=notes)

# Route pour modifier une note
@app.route('/modifier/<int:id>', methods=['GET', 'POST'])
def modifier_note(id):
    note = Note.query.get_or_404(id)  # Récupère la note à modifier par son ID
    form = NoteForm(obj=note)  # Remplir le formulaire avec les valeurs de la note existante

    if form.validate_on_submit():  # Si le formulaire est soumis et validé
        # Mise à jour des valeurs de la note
        note.titre = form.titre.data
        note.note = form.note.data

        # Sauvegarde des modifications dans la base de données
        db.session.commit()
        flash("Note modifiée avec succès", 'success')
        return redirect(url_for('gestion_notes'))  # Rediriger vers la page de gestion des notes

    # Si c'est une requête GET, affichez le formulaire avec les valeurs actuelles
    return render_template('modifier_note.html', form=form, note=note)

# Route pour supprimer une note
@app.route('/supprimer/<int:id>')
def supprimer_note(id):
    note = Note.query.get_or_404(id)
    db.session.delete(note)
    db.session.commit()
    flash("Note supprimée", 'danger')
    return redirect(url_for('gestion_notes'))

# Route pour exporter les notes vers Excel
@app.route('/exporter_excel')
def exporter_excel():
    notes = Note.query.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Titre', 'Note'])

    for note in notes:
        ws.append([note.titre, note.note])

    wb.save("notes.xlsx")
    flash("Fichier Excel exporté avec succès", 'success')
    return redirect(url_for('gestion_notes'))

# Route pour générer un PDF avec les notes
@app.route('/imprimer_pdf')
def imprimer_pdf():
    notes = Note.query.all()

    # Créer un PDF avec fpdf
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Titre du PDF
    pdf.set_font("Arial", style="B", size=16)
    pdf.cell(0, 10, "Liste des Notes", ln=True, align='C')
    pdf.ln(10)  # Espacement vertical

    # Contenu du PDF
    pdf.set_font("Arial", size=12)
    for note in notes:
        pdf.cell(0, 10, f"Titre : {note.titre} - Note : {note.note}", ln=True)

    # Sauvegarder le fichier PDF
    pdf_file = "notes.pdf"
    pdf.output(pdf_file)
    flash("Fichier PDF généré avec succès", 'success')
    return redirect(url_for('gestion_notes'))

if __name__ == '__main__':
    # Créer les tables dans la base de données si elles n'existent pas
    with app.app_context():
        db.create_all()
    app.run(debug=True)
